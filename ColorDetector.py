from openpyxl.worksheet.worksheet import Worksheet
from dotenv import load_dotenv
import os
import sys

# 설정파일 로드
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS        # PyInstaller 임시 폴더
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # 파이참 실행 시 현재 파일 경로
ENV_PATH = os.path.join(BASE_DIR, ".env")
load_dotenv(ENV_PATH)

# 설정파일에서 상수 설정
REPORT_DATE: int = int(os.getenv("REPORT_DATE"))
TRACKING_NUMBER: int = int(os.getenv("TRACKING_NUMBER"))
RESENT_NUMBER:int = int(os.getenv("RESENT_NUMBER"))
DELIVERY_DATE: int = int(os.getenv("DELIVERY_DATE"))
UNDELIVERED_REASON_DETAIL : int = int(os.getenv("UNDELIVERED_REASON_DETAIL"))

CLIENT_NAME : str = os.getenv("CLIENT_NAME")

class ColorDetector:
    @staticmethod
    def is_white(work_sheet: Worksheet, row_number: int) -> bool:
        ret = False

        # 1. 셀을 읽는다.
        cell = work_sheet.cell(row=row_number, column=REPORT_DATE)
        fg_color = cell.fill.fgColor

        color = None
        # 2. 배경색이 직접 RGB 값이면,
        if fg_color.type == "rgb" and fg_color.rgb is not None:
            color = fg_color.rgb

        # 3. 배경색이 테마 기반 색상이면,
        if fg_color.type == "theme":
            color = f"theme:{fg_color.theme}, tint:{fg_color.tint}"

        # 4. 배경색이 인덱스 기반이면,
        if fg_color.type == "indexed":
            color = f"indexed:{fg_color.indexed}"

        # 5. 배경색이 조건부 서식이면,
        # 5.1. 재발송 번호가 폐기이면,
        resent_number = work_sheet.cell(row=row_number, column=RESENT_NUMBER)
        if resent_number.value == "폐기":
            color = "conditionalStyle_29"
        # 5.2. 배송 완료일이 있으면,
        delivery_date = work_sheet.cell(row=row_number, column=DELIVERY_DATE)
        if delivery_date.value is not None:
            color = "conditionalStyle_28"
        # 5.3. 셀값이 30이면,
        if  cell.value == 30:
            color = "conditionalStyle_27"

        # 흰색이면, 참으로 설정한다.
        if color is None or color == "00000000":
            ret = True

        return ret

    @staticmethod
    def is_pink(work_sheet: Worksheet, row_number: int) -> bool:
        ret = False

        # 1. 셀을 읽는다.
        cell = work_sheet.cell(row=row_number, column=UNDELIVERED_REASON_DETAIL)

        # 2. 미배송 사유에 고객사명이 포함되어 있는지 확인한다.
        if cell.value is not None and CLIENT_NAME in cell.value:
            ret = True

        return ret
