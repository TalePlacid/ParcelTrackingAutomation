from openpyxl import load_workbook
from ParcelSheet import ParcelSheet
from dotenv import load_dotenv
import os
import sys
from ColorDetector import ColorDetector
import re
from openpyxl.styles import Font, Alignment, Border, Side

# 설정파일 로드
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS        # PyInstaller 임시 폴더
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # 파이참 실행 시 현재 파일 경로
ENV_PATH = os.path.join(BASE_DIR, ".env")
load_dotenv(ENV_PATH)

# 설정파일에서 상수 설정
REPORT_DATE: int = int(os.getenv("REPORT_DATE"))
TRACKING_NUMBER: int = int(os.getenv("TRACKING_NUMBER"))
RESENT_NUMBER:int = int(os.getenv("RESENT_NUMBER"))
DELIVERY_DATE: int = int(os.getenv("DELIVERY_DATE"))

DELIVERY_DATE_FORMAT: str = os.getenv("DELIVERY_DATE_FORMAT")
FONT_NAME: str = os.getenv("FONT_NAME")

class ExcelAdaptor:
    def __init__(self, path):
        self.path = path

        self.work_book = load_workbook(self.path) # 엑셀파일 열기
        self.work_book.iso_dates = False # 날짜형을 YYYY-MM-DD가 아닌 정수형(엑셀 기본)으로 사용
        self.work_sheet = self.work_book.worksheets[0] # 엑셀파일의 첫번째 시트 열기

    def load(self, parcel_sheet: ParcelSheet)->int:
        # 1. 표 헤더가 나올 때까지 내려간다.
        i = 1
        while self.work_sheet.cell(row = i, column = REPORT_DATE).value != "보고날짜":
            i += 1

        # 2. 표의 끝이 아니면 반복한다.
        count = 0
        i += 1
        while self.work_sheet.cell(row = i, column = REPORT_DATE).value is not None:
            # 2.1. 색이 흰색이거나, 분홍색이면,
            if ColorDetector.is_white(self.work_sheet, i) or ColorDetector.is_pink(self.work_sheet, i):
                # 2.1.1. 재배송번호를 읽는다.
                tracking_number = str(self.work_sheet.cell(row = i, column = RESENT_NUMBER).value)
                is_resent = True

                # 2.1.2. 재배송번호가 유효하지 않으면, 송장번호를 읽는다.
                tracking_number_is_valid = bool(re.fullmatch(r"\d{12}", tracking_number)) # 숫자 12자리
                if not tracking_number_is_valid:
                    tracking_number = str(self.work_sheet.cell(row = i, column = TRACKING_NUMBER).value)
                    is_resent = False

                # 2.1.3. 택배기록표에 기재한다.
                parcel_sheet.record_from_excel_row(int(tracking_number), is_resent, i)
                count += 1
            i += 1

        return count

    def update(self, parcel_sheet: ParcelSheet)->int:
        # 1. 배송기록표의 처음부터 끝까지 반복한다.
        thin = Side(border_style="thin", color="000000")  # 실선, 검은색
        border = Border(left=thin, right=thin, top=thin, bottom=thin) # 테두리 설정
        count = 0
        i = 0
        while i < parcel_sheet.length:
            # 1.1. 배송완료일이 있다면,
            parcel = parcel_sheet[i]
            if parcel.delivery_date is not None:
                # 1.1.1. 행에 해당하는 셀을 읽는다.
                cell = self.work_sheet.cell(row = parcel.row_number, column = DELIVERY_DATE)

                # 1.1.2. 배송완료일을 반영한다.
                cell.value = parcel.delivery_date
                cell.number_format = DELIVERY_DATE_FORMAT
                cell.font = Font(name = FONT_NAME, size=10, bold=False)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

                count += 1
            i += 1

        # 2. 엑셀파일을 저장한다.
        self.work_book.save(self.path)

        return count